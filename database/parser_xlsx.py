#!/usr/bin/env python3
# -*- coding=utf-8 -*-
from openpyxl import load_workbook
import re


def excel_parser_l_d(file_name='access.xlsx', sheet_name='sheet1'):
    data = load_workbook(file_name)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    port_num_l = []
    e_device_l = []
    for row in table.iter_rows():
        # 分析被迭代出来的这个row的第2号位（即第C列）的值是否是我们想要的。
        if re.match(r'ZGW-JR\d+-\d+M', row[2].value):
            port_num_l.append(int(re.match(r'ZGW-JR\d+-(\d+)M', row[2].value).groups()[0]))
            # 同时提取第4号位的值
            e_device_l.append(row[4].value)
        else:
            pass
    # 返回列表
    # return port_num_l, e_device_l
    # 返回字典
    return dict(zip(port_num_l, e_device_l))
    # print(port_num_l)
    # print(e_device_l)


if __name__ == "__main__":
    print(excel_parser_l_d('access.xlsx', '1'))
